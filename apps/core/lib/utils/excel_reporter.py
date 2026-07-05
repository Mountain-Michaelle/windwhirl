
import logging
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelReporter:
    """
    Generates a status Excel report from the original customers.xlsx.

    Takes the original Excel file as a base, adds one column
    "Status/Comment" after the WhatsApp Number column, fills it
    with the send result for each customer from the database,
    and saves to reports/send_report_YYYY-MM-DD.xlsx.

    Usage:
        reporter = ExcelReporter(cfg)
        path = reporter.generate(db)
        # path → Path("reports/send_report_2025-06-28.xlsx")
    """

    # ── Cell colours for visual scanning ──────────────────────
    # Green fill for sent, red fill for not sent
    # Makes the report scannable at a glance without reading
    GREEN_FILL = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )
    RED_FILL = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )
    HEADER_FILL = PatternFill(
        start_color="1B4F8A",
        end_color="1B4F8A",
        fill_type="solid"
    )

    def __init__(self, cfg):
        """
        Args:
            cfg: AppConfig instance — provides excel_path() and country_code
        """
        self._cfg = cfg
        self._log = logging.getLogger(self.__class__.__name__)

    def generate(self, db) -> Path:
        """
        Build the status Excel report and save it to reports/.

        Flow:
          1. Read original customers.xlsx into pandas
          2. Pull all send statuses from database
          3. Match each Excel row to its DB status by Order ID
          4. Build "Status/Comment" value per row
          5. Insert new column after WhatsApp Number
          6. Apply colour coding (green=sent, red=not sent)
          7. Save to reports/send_report_YYYY-MM-DD.xlsx

        Args:
            db: Database instance

        Returns:
            Path to the saved Excel file.

        Raises:
            FileNotFoundError if original Excel is missing.
        """
        excel_path = self._cfg.excel_path()
        if not excel_path.exists():
            raise FileNotFoundError(
                f"Original Excel not found: {excel_path}\n"
                f"Cannot generate report without source file."
            )

        self._log.info(f"Building Excel report from: {excel_path}")

        # ── Step 1: Read original Excel ────────────────────────
        # Keep all columns exactly as they are in the original.
        # dtype=str prevents pandas from mangling phone numbers.
        df = pd.read_excel(excel_path, dtype=str, engine="openpyxl")

        # ── Step 2: Get all send statuses from DB ──────────────
        # Returns dict: { order_id: {"status": ..., "error": ...} }
        status_map = db.get_all_statuses()

        # ── Step 3: Build Status/Comment value for each row ────
        status_comments = []

        for _, row in df.iterrows():
            order_id = str(row.get("Order ID", "") or "").strip()

            if not order_id or order_id.lower() == "nan":
                # Row has no Order ID — cannot match to DB
                status_comments.append("False / Not sent — No Order ID found")
                continue

            record = status_map.get(order_id)

            if record is None:
                # Order ID exists in Excel but not in database
                # (row was skipped during import — wrong product etc.)
                status_comments.append(
                    "False / Not sent — Not in target product filter"
                )
                continue

            status        = record["status"]
            error_message = record.get("error_message", "") or ""

            # ── Map DB status to human-readable Status/Comment ─
            if status == "SENT":
                status_comments.append("True / Sent")

            elif status == "INVALID_NUMBER":
                status_comments.append(
                    "False / Not sent — Phone not registered on WhatsApp"
                )

            elif status == "FAILED_FINAL":
                reason = error_message if error_message else "Send failed after 2 attempts"
                status_comments.append(f"False / Not sent — {reason}")

            elif status == "FAILED":
                reason = error_message if error_message else "Send failed — eligible for retry"
                status_comments.append(f"False / Not sent — {reason}")

            elif status == "PENDING":
                status_comments.append(
                    "False / Not sent — Pending (not yet attempted)"
                )

            elif status == "INVALID_PHONE":
                status_comments.append(
                    "False / Not sent — Phone number could not be normalized"
                )

            else:
                # Catch-all for any unexpected status value
                status_comments.append(
                    f"False / Not sent — {status}"
                )

        # ── Step 4: Insert Status/Comment column ───────────────
        # Find position of WhatsApp Number column
        # Insert our new column immediately after it
        columns = list(df.columns)

        if "WhatsApp Number" in columns:
            insert_position = columns.index("WhatsApp Number") + 1
        else:
            # WhatsApp Number column not found — append at end
            self._log.warning(
                "WhatsApp Number column not found in Excel. "
                "Status/Comment will be appended at the end."
            )
            insert_position = len(columns)

        # Insert the new column into the dataframe
        df.insert(
            loc=insert_position,
            column="Status/Comment",
            value=status_comments
        )

        # ── Step 5: Save to reports/ ───────────────────────────
        Path("reports").mkdir(exist_ok=True)
        today_str   = date.today().strftime("%Y-%m-%d")
        output_path = Path("reports") / f"send_report_{today_str}.xlsx"

        # Save via pandas first (handles data correctly)
        df.to_excel(output_path, index=False, engine="openpyxl")

        # ── Step 6: Apply colour coding with openpyxl ──────────
        # Re-open the saved file to apply styling
        # pandas doesn't support cell-level styling directly
        self._apply_styling(output_path, insert_position + 1)
        # +1 because openpyxl columns are 1-indexed

        self._log.info(
            f"Excel report saved: {output_path} "
            f"({len(df)} rows, {len(df.columns)} columns)"
        )

        return output_path

    def _apply_styling(self, file_path: Path, status_col_index: int):
        """
        Apply visual styling to the saved Excel file.

        - Header row: dark blue background, white bold text
        - Status/Comment column header: same dark blue
        - "True / Sent" cells: green background
        - "False / Not sent" cells: red background
        - Auto-width on Status/Comment column

        Args:
            file_path:        Path to the Excel file to style.
            status_col_index: 1-based column index of Status/Comment.
        """
        try:
            wb   = load_workbook(file_path)
            ws   = wb.active
            col  = status_col_index
            col_letter = get_column_letter(col)

            # Style the header row (row 1)
            for cell in ws[1]:
                cell.fill      = self.HEADER_FILL
                cell.font      = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

            # Style each data cell in Status/Comment column
            for row_num in range(2, ws.max_row + 1):
                cell  = ws[f"{col_letter}{row_num}"]
                value = str(cell.value or "")

                if value.startswith("True"):
                    cell.fill = self.GREEN_FILL
                elif value.startswith("False"):
                    cell.fill = self.RED_FILL

                # Wrap text so long error messages are readable
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )

            # Auto-width the Status/Comment column
            # (set to 45 chars — enough for most error messages)
            ws.column_dimensions[col_letter].width = 45

            # Freeze the header row so it stays visible when scrolling
            ws.freeze_panes = "A2"

            wb.save(file_path)
            self._log.debug(f"Styling applied to: {file_path}")

        except Exception as e:
            # Styling failure must not prevent the report from saving
            # The data is already saved — styling is cosmetic only
            self._log.warning(f"Could not apply Excel styling: {e}")
