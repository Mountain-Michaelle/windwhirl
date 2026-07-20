from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Optional

from apps.oms.infrastructure.persistence.order_repository import OrderRepository
from apps.oms.infrastructure.persistence.schema import OrderRecord
from apps.oms.shared.logger import get_logger

log = get_logger(__name__)


class ExcelExporter:
    '''
    Exports order data from the database to Excel files.

    Uses openpyxl directly — no pandas dependency for this module.
    Each exported file goes to the reports/ directory.

    Usage:
        exporter = ExcelExporter(order_repo, reports_dir="reports")
        path = await exporter.export_daily()
        path = await exporter.export_by_worker("2348XXXXXXXXX")
    '''

    # Column definitions: (header_label, record_attribute, column_width)
    COLUMNS = [
        ("Order ID",          "order_id",          15),
        ("Customer Name",     "customer_name",      20),
        ("Phone Number",      "phone_number",       16),
        ("WhatsApp Number",   "whatsapp_number",    16),
        ("Package",           "package_name",       20),
        ("Price",             "price_raw",          12),
        ("Delivery Address",  "delivery_address",   35),
        ("Delivery Request",  "delivery_request",   18),
        ("Order Date",        "order_date_raw",     14),
        ("Campaign",          "campaign",           16),
        ("Customer Question", "customer_question",  25),
        ("Worker",            "worker_number",      16),
        ("Assignment Status", "assignment_status",  18),
        ("Duplicate Status",  "duplicate_status",   18),
        ("Quality Score",     "quality_score",      14),
        ("Valid",             "is_valid",           8),
        ("Created At",        "created_at",         18),
        ("Assigned At",       "assigned_at",        18),
    ]

    def __init__(
        self,
        order_repo:  OrderRepository,
        reports_dir: str = "reports",
    ):
        self._repo        = order_repo
        self._reports_dir = Path(reports_dir)
        self._reports_dir.mkdir(parents=True, exist_ok=True)

    async def export_daily(
        self,
        export_date:   date = None,
        worker_number: str  = None,
    ) -> Path:
        '''
        Export all orders for a given date.
        Optionally filtered by worker.

        Args:
            export_date:   Date to export. Defaults to today.
            worker_number: Optional worker filter.

        Returns:
            Path to the generated Excel file.
        '''
        target_date   = export_date or date.today()
        records       = await self._repo.get_today(worker_number)

        date_str  = target_date.strftime("%Y-%m-%d")
        worker_str = f"_{worker_number}" if worker_number else ""
        filename   = f"orders_{date_str}{worker_str}.xlsx"
        path       = self._reports_dir / filename

        self._write_excel(records, path, title=f"Orders — {date_str}")
        log.info(f"ExcelExporter: exported {len(records)} order(s) to {path}")
        return path

    async def export_by_worker(
        self,
        worker_number: str,
        status:        str = None,
    ) -> Path:
        '''
        Export all orders assigned to a worker.
        Optionally filtered by assignment_status.

        Args:
            worker_number: Worker phone number.
            status:        Optional status filter e.g. "ASSIGNED".

        Returns:
            Path to the generated Excel file.
        '''
        records   = await self._repo.get_by_worker(worker_number, status)
        date_str  = date.today().strftime("%Y-%m-%d")
        safe_num  = worker_number.replace("+", "")
        filename  = f"worker_{safe_num}_{date_str}.xlsx"
        path      = self._reports_dir / filename

        self._write_excel(records, path, title=f"Orders — +{worker_number}")
        log.info(
            f"ExcelExporter: exported {len(records)} order(s) "
            f"for +{worker_number} to {path}"
        )
        return path

    def _write_excel(
        self,
        records: list[OrderRecord],
        path:    Path,
        title:   str = "Orders",
    ) -> None:
        '''
        Write records to an Excel file using openpyxl.

        Args:
            records: List of OrderRecord objects.
            path:    Output file path.
            title:   Worksheet title.
        '''
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel limit: 31 chars

        # Header fills
        HEADER_FILL = PatternFill("solid", fgColor="1B4F8A")
        VALID_FILL  = PatternFill("solid", fgColor="C6EFCE")
        INVALID_FILL= PatternFill("solid", fgColor="FFC7CE")
        DUP_FILL    = PatternFill("solid", fgColor="FFEB9C")

        # Write header row
        for col_idx, (header, _, width) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill      = HEADER_FILL
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            ws.column_dimensions[
                __import__('openpyxl').utils.get_column_letter(col_idx)
            ].width = width

        # Write data rows
        for row_idx, record in enumerate(records, 2):
            for col_idx, (_, attr, _) in enumerate(self.COLUMNS, 1):
                value = getattr(record, attr, None)

                # Format datetimes
                if isinstance(value, datetime):
                    value = value.strftime("%d/%m/%Y %H:%M")

                # Format booleans
                if isinstance(value, bool):
                    value = "Yes" if value else "No"

                # Format floats (quality score)
                if isinstance(value, float) and attr == "quality_score":
                    value = f"{value:.0%}"

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Row highlighting
            is_valid   = getattr(record, 'is_valid', True)
            dup_status = getattr(record, 'duplicate_status', 'UNIQUE')

            if dup_status in ("CONFIRMED_DUPLICATE", "LIKELY_DUPLICATE"):
                fill = DUP_FILL
            elif is_valid:
                fill = VALID_FILL
            else:
                fill = INVALID_FILL

            for col_idx in range(1, len(self.COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter on header row
        ws.auto_filter.ref = ws.dimensions

        wb.save(path)
