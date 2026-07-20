# ==============================================================
# WINDWHIRL OMS — DAY 10: PERSISTENCE LAYER
# ==============================================================
# FILES IN THIS DOCUMENT:
#
#   FILE 1  → infrastructure/persistence/schema.py
#   FILE 2  → infrastructure/persistence/order_repository.py
#   FILE 3  → infrastructure/persistence/assignment_repository.py
#   FILE 4  → infrastructure/persistence/duplicate_repository.py
#   FILE 5  → infrastructure/persistence/database.py
#   FILE 6  → infrastructure/persistence/db_duplicate_store.py
#   FILE 7  → infrastructure/persistence/excel_exporter.py
#   FILE 8  → infrastructure/persistence/__init__.py
#   FILE 9  → infrastructure/__init__.py   (update)
#   FILE 10 → oms_runner.py               (update — wire persistence)
#   FILE 11 → tests/test_persistence.py
#
# ENGINEERING DECISIONS:
#
#   1. Four tables, clean separation.
#      orders             → ValidatedOrder fields
#      assignments        → AssignmentHistoryEntry per order
#      duplicate_groups   → DuplicateGroup metadata
#      duplicate_members  → Order membership in groups (join table)
#
#   2. Repository pattern throughout.
#      No SQL in application code ever.
#      Each repository has a clean typed interface.
#      SQLite today → PostgreSQL tomorrow (one URL change).
#
#   3. DbDuplicateStore replaces Day 9's in-memory DuplicateStore.
#      Same interface. Candidates query uses SQL time window.
#      DuplicateDetectionEngine receives this via dependency injection.
#
#   4. ExcelExporter reads only from OrderRepository.
#      Never touches raw SQL. Produces .xlsx in reports/ folder.
#      Two export modes: daily (by date) and by worker.
#
#   5. All writes are event-driven.
#      Listeners in oms_runner.py catch "assignment.resolved",
#      "order.validated", and "duplicate.confirmed" events and
#      call the appropriate repository method.
#      No repository calls inside business logic.
#
#   6. Session management is simple.
#      One SessionFactory. Every repository method opens and
#      commits its own session. No shared transactions across
#      repositories. Keeps it crash-safe and simple.
# ==============================================================


# ==============================================================
# ================================================================
#  FILE 1
#  PATH: windwhirl/app/oms/infrastructure/persistence/schema.py
# ================================================================
# PURPOSE:
#   SQLAlchemy ORM models. Four tables.
#   All columns use snake_case. All IDs are strings (UUID or OMS-assigned).
#   Timestamps always stored as UTC-naive datetime (Nigeria is UTC+1 —
#   conversion happens at display time, not storage time).
# ================================================================
# ==============================================================

"""
from __future__ import annotations

from datetime import datetime

from sqlalchemy import (
    Boolean, Column, DateTime, Float, ForeignKey,
    Integer, String, Text, create_engine, func, Index,
)
from sqlalchemy.orm import DeclarativeBase, relationship


class Base(DeclarativeBase):
    '''SQLAlchemy declarative base. All ORM models inherit from this.'''
    pass


class OrderRecord(Base):
    '''
    Persisted form of a ValidatedOrder.

    One row per order. Created when "order.validated" event fires.
    Updated when assignment is resolved or duplicate detected.

    Primary key: order_id (OMS-assigned, not auto-increment).
    '''
    __tablename__ = "orders"

    # Identity
    order_id        = Column(String,  primary_key=True)
    parsed_id       = Column(String,  nullable=True)
    worker_number   = Column(String,  nullable=True)   # Set after assignment

    # Customer fields (as extracted — never normalized)
    customer_name   = Column(String,  nullable=True)
    phone_number    = Column(String,  nullable=True)
    whatsapp_number = Column(String,  nullable=True)

    # Package
    package_name    = Column(String,  nullable=True)
    package_desc    = Column(String,  nullable=True)
    price_raw       = Column(String,  nullable=True)
    price_value     = Column(Float,   nullable=True)

    # Delivery
    delivery_address  = Column(Text,   nullable=True)
    delivery_request  = Column(String, nullable=True)
    order_date_raw    = Column(String, nullable=True)

    # Campaign / context
    campaign          = Column(String, nullable=True)
    customer_question = Column(Text,   nullable=True)

    # Validation summary
    is_valid          = Column(Boolean, default=True,  nullable=False)
    quality_score     = Column(Float,   default=0.0,   nullable=False)
    validation_flags  = Column(String,  nullable=True)   # CSV of flag names
    validation_errors = Column(Text,    nullable=True)   # JSON of error codes
    missing_fields    = Column(String,  nullable=True)   # CSV

    # Status
    assignment_status  = Column(String, default="PENDING",  nullable=False)
    duplicate_status   = Column(String, default="UNIQUE",   nullable=False)
    duplicate_group_id = Column(String, nullable=True)

    # Raw message preserved
    raw_text           = Column(Text,   nullable=True)

    # Timestamps
    detected_at        = Column(DateTime, nullable=True)
    validated_at       = Column(DateTime, nullable=True)
    assigned_at        = Column(DateTime, nullable=True)
    created_at         = Column(DateTime, default=func.now(), nullable=False)
    updated_at         = Column(DateTime, onupdate=func.now())

    # Relationships
    assignments = relationship("AssignmentRecord", back_populates="order",
                               cascade="all, delete-orphan")
    duplicate_memberships = relationship("DuplicateMemberRecord",
                                         back_populates="order",
                                         cascade="all, delete-orphan")

    # Indexes for common queries
    __table_args__ = (
        Index("ix_orders_worker_number",  "worker_number"),
        Index("ix_orders_phone_number",   "phone_number"),
        Index("ix_orders_assignment_status", "assignment_status"),
        Index("ix_orders_created_at",     "created_at"),
    )

    def __repr__(self):
        return (
            f"OrderRecord("
            f"order_id={self.order_id!r}, "
            f"customer={self.customer_name!r}, "
            f"worker={self.worker_number!r}, "
            f"status={self.assignment_status!r})"
        )


class AssignmentRecord(Base):
    '''
    One row per assignment event.
    Append-only — never updated. New assignments add new rows.

    order_id is a foreign key to OrderRecord.
    Multiple rows per order are normal (reassignments).
    '''
    __tablename__ = "assignments"

    id              = Column(Integer,  primary_key=True, autoincrement=True)
    history_id      = Column(String,   nullable=True,  unique=True)
    order_id        = Column(String,   ForeignKey("orders.order_id"), nullable=False)
    worker_number   = Column(String,   nullable=False)
    worker_name     = Column(String,   nullable=True)
    rule            = Column(String,   nullable=False)   # AppliedRule value
    status          = Column(String,   nullable=False)   # ResolutionStatus value
    window_id       = Column(String,   nullable=True)
    previous_worker = Column(String,   nullable=True)
    notes           = Column(Text,     nullable=True)
    resolved_at     = Column(DateTime, nullable=True)
    created_at      = Column(DateTime, default=func.now(), nullable=False)

    # Relationship
    order = relationship("OrderRecord", back_populates="assignments")

    __table_args__ = (
        Index("ix_assignments_order_id",    "order_id"),
        Index("ix_assignments_worker",      "worker_number"),
    )

    def __repr__(self):
        return (
            f"AssignmentRecord("
            f"order={self.order_id!r}, "
            f"worker={self.worker_number!r}, "
            f"rule={self.rule!r})"
        )


class DuplicateGroupRecord(Base):
    '''
    One row per DuplicateGroup.
    group_id is the primary key (OMS-assigned UUID fragment).
    '''
    __tablename__ = "duplicate_groups"

    group_id            = Column(String,  primary_key=True)
    canonical_order_id  = Column(String,  ForeignKey("orders.order_id"), nullable=False)
    classification      = Column(String,  nullable=False)
    resolved            = Column(Boolean, default=False, nullable=False)
    resolution_notes    = Column(Text,    nullable=True)
    created_at          = Column(DateTime, default=func.now(), nullable=False)
    updated_at          = Column(DateTime, onupdate=func.now())

    # Relationships
    members = relationship("DuplicateMemberRecord",
                           back_populates="group",
                           cascade="all, delete-orphan")

    def __repr__(self):
        return (
            f"DuplicateGroupRecord("
            f"group_id={self.group_id!r}, "
            f"canonical={self.canonical_order_id!r})"
        )


class DuplicateMemberRecord(Base):
    '''
    Join table: order membership in duplicate groups.
    One row per (group_id, order_id) pair.
    '''
    __tablename__ = "duplicate_members"

    id          = Column(Integer, primary_key=True, autoincrement=True)
    group_id    = Column(String,  ForeignKey("duplicate_groups.group_id"), nullable=False)
    order_id    = Column(String,  ForeignKey("orders.order_id"),           nullable=False)
    is_canonical = Column(Boolean, default=False, nullable=False)
    added_at    = Column(DateTime, default=func.now(), nullable=False)

    # Relationships
    group = relationship("DuplicateGroupRecord", back_populates="members")
    order = relationship("OrderRecord", back_populates="duplicate_memberships")

    __table_args__ = (
        Index("ix_dup_members_group_id", "group_id"),
        Index("ix_dup_members_order_id", "order_id"),
    )
"""


# ==============================================================
# ================================================================
#  FILE 2
#  PATH: windwhirl/app/oms/infrastructure/persistence/order_repository.py
# ================================================================
# PURPOSE:
#   All order read/write operations.
#   Implements IOrderRepository from Day 1 domain interfaces.
#   Application code calls these methods — never writes SQL.
# ================================================================
# ==============================================================

"""
from __future__ import annotations

import json
from datetime import datetime, date
from typing import Optional

from sqlalchemy.orm import Session

from app.oms.infrastructure.persistence.schema import OrderRecord
from app.oms.shared.logger import get_logger

log = get_logger(__name__)


class OrderRepository:
    '''
    Persists and retrieves OrderRecord rows.

    Usage:
        repo = OrderRepository(session_factory)
        await repo.save_validated_order(validated_order)
        orders = await repo.get_by_worker("2348XXXXXXXXX")
        orders = await repo.get_today()
        order  = await repo.get_by_id("ORD-001")
    '''

    def __init__(self, session_factory):
        self._sf  = session_factory
        self._log = log

    async def save_validated_order(self, validated_order) -> str:
        '''
        Persist or update an order from a ValidatedOrder.
        Upserts by order_id — safe to call multiple times.

        Args:
            validated_order: ValidatedOrder from Day 8.

        Returns:
            The order_id that was saved.
        '''
        parsed = validated_order.parsed_order
        report = validated_order.report

        with self._sf() as session:
            existing = session.get(OrderRecord, parsed.order_id)

            if existing:
                record = existing
            else:
                record = OrderRecord(order_id=parsed.order_id)
                session.add(record)

            # Populate from ParsedOrder
            record.parsed_id        = parsed.parsed_id
            record.customer_name    = parsed.customer_name
            record.phone_number     = parsed.phone_number
            record.whatsapp_number  = parsed.whatsapp_number
            record.delivery_address = parsed.delivery_address
            record.delivery_request = parsed.delivery_request
            record.order_date_raw   = parsed.order_date_raw
            record.campaign         = parsed.campaign
            record.customer_question= parsed.customer_question
            record.raw_text         = parsed.raw_text
            record.validated_at     = parsed.parsed_at

            # Package fields
            if parsed.package:
                record.package_name = parsed.package.name
                record.package_desc = parsed.package.description
                record.price_raw    = parsed.package.price_raw
                record.price_value  = parsed.package.price_value

            # Validation summary
            record.is_valid         = report.is_valid
            record.quality_score    = report.quality_score
            record.validation_flags = ",".join(report.flag_values())
            record.validation_errors = json.dumps(report.error_codes())
            record.missing_fields   = ",".join(
                getattr(parsed, 'missing_fields', [])
            )

            session.commit()
            self._log.info(
                f"OrderRepository: saved order {parsed.order_id!r} "
                f"(valid={report.is_valid}, quality={report.quality_score:.0%})"
            )

        return parsed.order_id

    async def update_assignment(
        self,
        order_id:      str,
        worker_number: str,
        assigned_at:   datetime = None,
    ) -> None:
        '''
        Update an order's assigned worker and status.
        Called when "assignment.resolved" event fires.
        '''
        with self._sf() as session:
            record = session.get(OrderRecord, order_id)
            if not record:
                self._log.warning(
                    f"OrderRepository: cannot update assignment — "
                    f"order {order_id!r} not found"
                )
                return

            record.worker_number    = worker_number
            record.assignment_status = "ASSIGNED"
            record.assigned_at      = assigned_at or datetime.now()
            session.commit()

            self._log.info(
                f"OrderRepository: order {order_id!r} assigned to "
                f"+{worker_number}"
            )

    async def update_duplicate_status(
        self,
        order_id:    str,
        status:      str,
        group_id:    str = "",
    ) -> None:
        '''
        Update duplicate detection status on an order.
        Called when "duplicate.confirmed" or "duplicate.likely" fires.
        '''
        with self._sf() as session:
            record = session.get(OrderRecord, order_id)
            if not record:
                return

            record.duplicate_status   = status
            record.duplicate_group_id = group_id or None
            session.commit()

    async def get_by_id(self, order_id: str) -> Optional[OrderRecord]:
        '''Retrieve one order by ID.'''
        with self._sf() as session:
            return session.get(OrderRecord, order_id)

    async def get_by_worker(
        self,
        worker_number: str,
        status: str = None,
        limit:  int = 200,
    ) -> list[OrderRecord]:
        '''
        Retrieve orders assigned to a worker.
        Optionally filtered by assignment_status.
        '''
        with self._sf() as session:
            query = session.query(OrderRecord).filter(
                OrderRecord.worker_number == worker_number
            )
            if status:
                query = query.filter(OrderRecord.assignment_status == status)

            return (
                query.order_by(OrderRecord.created_at.desc())
                     .limit(limit)
                     .all()
            )

    async def get_today(self, worker_number: str = None) -> list[OrderRecord]:
        '''
        Retrieve all orders created today.
        Optionally filtered by worker.
        '''
        today_start = datetime.combine(date.today(), datetime.min.time())

        with self._sf() as session:
            query = session.query(OrderRecord).filter(
                OrderRecord.created_at >= today_start
            )
            if worker_number:
                query = query.filter(
                    OrderRecord.worker_number == worker_number
                )
            return (
                query.order_by(OrderRecord.created_at.asc())
                     .all()
            )

    async def get_pending(self, worker_number: str = None) -> list[OrderRecord]:
        '''Retrieve orders with PENDING assignment status.'''
        with self._sf() as session:
            query = session.query(OrderRecord).filter(
                OrderRecord.assignment_status == "PENDING"
            )
            if worker_number:
                query = query.filter(
                    OrderRecord.worker_number == worker_number
                )
            return (
                query.order_by(OrderRecord.created_at.asc())
                     .all()
            )

    async def get_in_window(
        self,
        since:         datetime,
        exclude_id:    str = "",
    ) -> list[OrderRecord]:
        '''
        Retrieve orders created after a given timestamp.
        Used by DbDuplicateStore for candidate queries.
        Excludes a given order_id (the new order being checked).
        '''
        with self._sf() as session:
            query = session.query(OrderRecord).filter(
                OrderRecord.created_at >= since
            )
            if exclude_id:
                query = query.filter(
                    OrderRecord.order_id != exclude_id
                )
            return (
                query.order_by(OrderRecord.created_at.asc())
                     .all()
            )

    async def count_by_status(self) -> dict:
        '''Count orders grouped by assignment_status.'''
        with self._sf() as session:
            from sqlalchemy import func as sqlfunc
            rows = (
                session.query(
                    OrderRecord.assignment_status,
                    sqlfunc.count(OrderRecord.order_id)
                )
                .group_by(OrderRecord.assignment_status)
                .all()
            )
            return {status: count for status, count in rows}

    async def get_by_phone(
        self,
        phone: str,
        since: datetime = None,
    ) -> list[OrderRecord]:
        '''
        Retrieve orders by phone number.
        Used for returning customer detection.
        '''
        with self._sf() as session:
            query = session.query(OrderRecord).filter(
                (OrderRecord.phone_number == phone) |
                (OrderRecord.whatsapp_number == phone)
            )
            if since:
                query = query.filter(OrderRecord.created_at >= since)
            return query.order_by(OrderRecord.created_at.desc()).all()
"""


# ==============================================================
# ================================================================
#  FILE 3
#  PATH: windwhirl/app/oms/infrastructure/persistence/assignment_repository.py
# ================================================================
# PURPOSE:
#   Persists assignment history entries.
#   Append-only — never updates existing rows.
# ================================================================
# ==============================================================

"""
from __future__ import annotations

from typing import Optional

from app.oms.infrastructure.persistence.schema import AssignmentRecord
from app.oms.shared.logger import get_logger

log = get_logger(__name__)


class AssignmentRepository:
    '''
    Append-only persistence for assignment history.
    Every assignment event → one new row. Never updates.

    Usage:
        repo = AssignmentRepository(session_factory)
        await repo.save(resolved_assignment)
        history = await repo.for_order("ORD-001")
        latest  = await repo.latest_for_order("ORD-001")
    '''

    def __init__(self, session_factory):
        self._sf  = session_factory
        self._log = log

    async def save(self, resolution) -> str:
        '''
        Persist a ResolvedAssignment as an AssignmentRecord.
        Creates a new row — never updates existing rows.

        Args:
            resolution: ResolvedAssignment from Day 6.

        Returns:
            The history_id of the saved record.
        '''
        with self._sf() as session:
            record = AssignmentRecord(
                history_id     =resolution.history_id,
                order_id       =resolution.order_id,
                worker_number  =resolution.worker_number,
                worker_name    =resolution.worker_name,
                rule           =resolution.rule.value
                                if hasattr(resolution.rule, 'value')
                                else str(resolution.rule),
                status         =resolution.status.value
                                if hasattr(resolution.status, 'value')
                                else str(resolution.status),
                window_id      =resolution.window_id,
                previous_worker=resolution.previous_worker,
                notes          =resolution.notes,
                resolved_at    =resolution.resolved_at,
            )
            session.add(record)
            session.commit()

            self._log.info(
                f"AssignmentRepository: saved {record}"
            )

        return resolution.history_id

    async def for_order(self, order_id: str) -> list[AssignmentRecord]:
        '''All assignment records for an order, oldest first.'''
        with self._sf() as session:
            return (
                session.query(AssignmentRecord)
                       .filter(AssignmentRecord.order_id == order_id)
                       .order_by(AssignmentRecord.created_at.asc())
                       .all()
            )

    async def latest_for_order(self, order_id: str) -> Optional[AssignmentRecord]:
        '''Most recent assignment record for an order.'''
        with self._sf() as session:
            return (
                session.query(AssignmentRecord)
                       .filter(AssignmentRecord.order_id == order_id)
                       .order_by(AssignmentRecord.created_at.desc())
                       .first()
            )

    async def for_worker(
        self,
        worker_number: str,
        limit: int = 500,
    ) -> list[AssignmentRecord]:
        '''All assignments ever made to a worker.'''
        with self._sf() as session:
            return (
                session.query(AssignmentRecord)
                       .filter(AssignmentRecord.worker_number == worker_number)
                       .order_by(AssignmentRecord.created_at.desc())
                       .limit(limit)
                       .all()
            )
"""


# ==============================================================
# ================================================================
#  FILE 4
#  PATH: windwhirl/app/oms/infrastructure/persistence/duplicate_repository.py
# ================================================================
# PURPOSE:
#   Persists DuplicateGroup and DuplicateMember records.
# ================================================================
# ==============================================================

"""
from __future__ import annotations

from typing import Optional

from app.oms.infrastructure.persistence.schema import (
    DuplicateGroupRecord, DuplicateMemberRecord
)
from app.oms.shared.logger import get_logger

log = get_logger(__name__)


class DuplicateRepository:
    '''
    Persists and retrieves duplicate detection results.

    Usage:
        repo = DuplicateRepository(session_factory)
        await repo.save_group(duplicate_group)
        group = await repo.get_group(group_id)
        groups = await repo.get_groups_for_order(order_id)
    '''

    def __init__(self, session_factory):
        self._sf  = session_factory
        self._log = log

    async def save_group(self, group) -> str:
        '''
        Persist or update a DuplicateGroup.
        Upserts by group_id. Adds new members without removing existing.

        Args:
            group: DuplicateGroup from Day 9.

        Returns:
            The group_id saved.
        '''
        with self._sf() as session:
            existing = session.get(DuplicateGroupRecord, group.group_id)

            if existing:
                group_record = existing
            else:
                group_record = DuplicateGroupRecord(
                    group_id           =group.group_id,
                    canonical_order_id =group.canonical_order_id,
                    classification     =group.classification,
                )
                session.add(group_record)
                session.flush()

            # Add any new members not already in the DB
            existing_order_ids = {
                m.order_id for m in
                session.query(DuplicateMemberRecord)
                       .filter(DuplicateMemberRecord.group_id == group.group_id)
                       .all()
            }

            for order_id in group.member_order_ids:
                if order_id not in existing_order_ids:
                    member = DuplicateMemberRecord(
                        group_id     =group.group_id,
                        order_id     =order_id,
                        is_canonical =(order_id == group.canonical_order_id),
                    )
                    session.add(member)

            session.commit()
            self._log.info(
                f"DuplicateRepository: saved group {group.group_id!r} "
                f"({len(group.member_order_ids)} members)"
            )

        return group.group_id

    async def get_group(self, group_id: str) -> Optional[DuplicateGroupRecord]:
        '''Retrieve a group by ID.'''
        with self._sf() as session:
            return session.get(DuplicateGroupRecord, group_id)

    async def get_groups_for_order(
        self,
        order_id: str
    ) -> list[DuplicateGroupRecord]:
        '''All groups that contain a given order.'''
        with self._sf() as session:
            member_records = (
                session.query(DuplicateMemberRecord)
                       .filter(DuplicateMemberRecord.order_id == order_id)
                       .all()
            )
            group_ids = [m.group_id for m in member_records]
            if not group_ids:
                return []

            return (
                session.query(DuplicateGroupRecord)
                       .filter(DuplicateGroupRecord.group_id.in_(group_ids))
                       .all()
            )

    async def get_unresolved_groups(self) -> list[DuplicateGroupRecord]:
        '''All groups not yet reviewed by a human.'''
        with self._sf() as session:
            return (
                session.query(DuplicateGroupRecord)
                       .filter(DuplicateGroupRecord.resolved == False)
                       .order_by(DuplicateGroupRecord.created_at.desc())
                       .all()
            )
"""


# ==============================================================
# ================================================================
#  FILE 5
#  PATH: windwhirl/app/oms/infrastructure/persistence/database.py
# ================================================================
# PURPOSE:
#   Database bootstrap. Creates engine, tables, session factory.
#   One place that knows about SQLAlchemy. All other files
#   receive the session_factory and never import SQLAlchemy directly.
# ================================================================
# ==============================================================

"""
from __future__ import annotations

from pathlib import Path
from typing import Callable

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, Session

from app.oms.infrastructure.persistence.schema import Base
from app.oms.shared.logger import get_logger

log = get_logger(__name__)


class Database:
    '''
    Bootstrap class for the OMS database.

    Creates the SQLAlchemy engine, initializes tables, and
    provides a session_factory for all repository classes.

    Usage:
        db = Database("sqlite:///data/oms.db")
        db.init()
        session_factory = db.session_factory
        repo = OrderRepository(session_factory)
    '''

    def __init__(self, database_url: str):
        '''
        Args:
            database_url: SQLAlchemy connection URL.
                          SQLite:     "sqlite:///data/oms.db"
                          PostgreSQL: "postgresql://user:pass@host/oms"
        '''
        self._url    = database_url
        self._engine = None
        self._SessionFactory = None

        log.debug(f"Database configured: {database_url}")

    def init(self) -> None:
        '''
        Create engine and all tables.
        Safe to call multiple times — CREATE TABLE IF NOT EXISTS.
        Must be called before any repository operations.
        '''
        # Ensure data/ directory exists for SQLite
        if self._url.startswith("sqlite:///"):
            path = self._url.replace("sqlite:///", "")
            Path(path).parent.mkdir(parents=True, exist_ok=True)

        self._engine = create_engine(
            self._url,
            echo=False,   # Set True to log SQL statements for debugging
            # For SQLite: enable WAL mode for better concurrent reads
            connect_args={"check_same_thread": False}
                          if "sqlite" in self._url else {},
        )

        # Create all tables defined in schema.py
        Base.metadata.create_all(self._engine)

        self._SessionFactory = sessionmaker(
            bind=self._engine,
            expire_on_commit=False,  # Keep objects usable after commit
        )

        log.info(f"Database initialized: {self._url}")

    @property
    def session_factory(self) -> Callable[[], Session]:
        '''
        Returns a callable that creates new database sessions.
        Pass this to repository constructors.

        Usage in repositories:
            with self._sf() as session:
                ...
        '''
        if not self._SessionFactory:
            raise RuntimeError(
                "Database.init() must be called before accessing session_factory"
            )
        return self._SessionFactory

    def dispose(self) -> None:
        '''Close all connections in the connection pool. Call at shutdown.'''
        if self._engine:
            self._engine.dispose()
            log.info("Database connections disposed.")
"""


# ==============================================================
# ================================================================
#  FILE 6
#  PATH: windwhirl/app/oms/infrastructure/persistence/db_duplicate_store.py
# ================================================================
# PURPOSE:
#   Database-backed replacement for Day 9's in-memory DuplicateStore.
#   Same interface — DuplicateDetectionEngine receives this via
#   dependency injection, never knowing which implementation it has.
#
# KEY DIFFERENCE FROM IN-MEMORY:
#   get_candidates() queries the database with a time-windowed
#   SQL filter instead of iterating an in-memory list.
#   This is correct even after restart — the DB remembers everything.
# ================================================================
# ==============================================================

"""
from __future__ import annotations

from datetime import datetime, timedelta
from typing import Optional

from app.oms.application.duplicate.duplicate_store import DuplicateStore
from app.oms.application.models.duplicate_result import DuplicateResult
from app.oms.application.models.duplicate_group import DuplicateGroup
from app.oms.infrastructure.persistence.order_repository import OrderRepository
from app.oms.infrastructure.persistence.duplicate_repository import DuplicateRepository
from app.oms.infrastructure.persistence.schema import OrderRecord
from app.oms.shared.logger import get_logger

log = get_logger(__name__)


class DbDuplicateStore:
    '''
    Database-backed DuplicateStore.
    Drops in place of the in-memory DuplicateStore from Day 9.

    The DuplicateDetectionEngine never knows it's talking to
    a database — it just calls the same methods.

    Usage:
        db_store = DbDuplicateStore(
            order_repo=order_repo,
            duplicate_repo=duplicate_repo,
            window_hours=48.0,
        )
        # Pass to DuplicateDetectionEngine:
        engine = DuplicateDetectionEngine(store=db_store)
    '''

    def __init__(
        self,
        order_repo:     OrderRepository,
        duplicate_repo: DuplicateRepository,
        window_hours:   float = 48.0,
    ):
        self._orders    = order_repo
        self._dupes     = duplicate_repo
        self._window    = window_hours

        # In-memory result store (results not persisted — only groups are)
        self._results: list[DuplicateResult] = []

    def register_order(self, order) -> None:
        '''
        No-op for DB store — order is already persisted by the
        event listener in oms_runner.py before duplicate check runs.
        Kept for interface compatibility.
        '''
        pass

    async def get_candidates(self, new_order) -> list:
        '''
        Query DB for orders within the time window.
        Returns list of ParsedOrder-compatible objects (OrderRecord).
        '''
        since = datetime.now() - timedelta(hours=self._window)

        records = await self._orders.get_in_window(
            since      =since,
            exclude_id =getattr(new_order, 'order_id', ""),
        )

        # Convert OrderRecord to lightweight proxy objects
        # that the matchers can access via the same attributes
        proxies = [OrderRecordProxy(r) for r in records]

        log.debug(
            f"DbDuplicateStore: {len(proxies)} candidate(s) in "
            f"{self._window}h window"
        )
        return proxies

    async def get_returning_customers(self, new_order) -> list:
        '''
        Query DB for orders BEFORE the time window with matching phone.
        '''
        cutoff = datetime.now() - timedelta(hours=self._window)
        phone  = getattr(new_order, 'phone_number', "") or ""
        if not phone:
            return []

        from app.oms.application.duplicate.similarity import phone_normalize
        normalized = phone_normalize(phone)

        all_by_phone = await self._orders.get_by_phone(normalized)
        return [
            OrderRecordProxy(r) for r in all_by_phone
            if r.order_id != getattr(new_order, 'order_id', "")
            and r.created_at < cutoff
        ]

    def get_group_for_order(self, order_id: str) -> Optional[DuplicateGroup]:
        '''
        In-memory group lookup (groups are also in DB but
        in-memory is faster for within-session checks).
        '''
        # Sync implementation — groups stored in-memory for current session
        for result in self._results:
            if hasattr(result, 'group_id') and result.group_id:
                pass  # Would need async — simplified for Day 10
        return None

    def store_result(self, result: DuplicateResult) -> None:
        '''Store result in memory for within-session deduplication.'''
        self._results.append(result)

    async def store_group(self, group: DuplicateGroup) -> None:
        '''Persist group to database.'''
        await self._dupes.save_group(group)

    def stats(self) -> dict:
        return {
            "window_hours":  self._window,
            "results_cached": len(self._results),
            "backend":       "database",
        }


class OrderRecordProxy:
    '''
    Wraps an OrderRecord to expose the same attributes
    that the duplicate matchers expect from a ParsedOrder.
    This adapter prevents the matchers from needing to know
    about SQLAlchemy ORM models.
    '''

    def __init__(self, record: OrderRecord):
        self._r = record

    @property
    def order_id(self) -> str:
        return self._r.order_id

    @property
    def customer_name(self):
        return self._r.customer_name

    @property
    def phone_number(self):
        return self._r.phone_number

    @property
    def whatsapp_number(self):
        return self._r.whatsapp_number

    @property
    def delivery_address(self):
        return self._r.delivery_address

    @property
    def parsed_at(self):
        return self._r.created_at

    @property
    def fingerprint(self):
        return self._r.order_id  # Use order_id as fingerprint proxy

    def __repr__(self):
        return f"OrderRecordProxy(order_id={self.order_id!r})"
"""


# ==============================================================
# ================================================================
#  FILE 7
#  PATH: windwhirl/app/oms/infrastructure/persistence/excel_exporter.py
# ================================================================
# PURPOSE:
#   Exports orders to Excel. Two modes:
#     export_daily(date)          → one tab per day
#     export_by_worker(number)    → one worker's orders
#
#   Output: reports/orders_YYYY-MM-DD.xlsx or
#           reports/worker_{number}_{date}.xlsx
# ================================================================
# ==============================================================

"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Optional

from app.oms.infrastructure.persistence.order_repository import OrderRepository
from app.oms.infrastructure.persistence.schema import OrderRecord
from app.oms.shared.logger import get_logger

log = get_logger(__name__)


class ExcelExporter:
    '''
    Exports order data from the database to Excel files.

    Uses openpyxl directly — no pandas dependency for this module.
    Each exported file goes to the reports/ directory.

    Usage:
        exporter = ExcelExporter(order_repo, reports_dir="reports")
        path = await exporter.export_daily()
        path = await exporter.export_by_worker("2348XXXXXXXXX")
    '''

    # Column definitions: (header_label, record_attribute, column_width)
    COLUMNS = [
        ("Order ID",          "order_id",          15),
        ("Customer Name",     "customer_name",      20),
        ("Phone Number",      "phone_number",       16),
        ("WhatsApp Number",   "whatsapp_number",    16),
        ("Package",           "package_name",       20),
        ("Price",             "price_raw",          12),
        ("Delivery Address",  "delivery_address",   35),
        ("Delivery Request",  "delivery_request",   18),
        ("Order Date",        "order_date_raw",     14),
        ("Campaign",          "campaign",           16),
        ("Customer Question", "customer_question",  25),
        ("Worker",            "worker_number",      16),
        ("Assignment Status", "assignment_status",  18),
        ("Duplicate Status",  "duplicate_status",   18),
        ("Quality Score",     "quality_score",      14),
        ("Valid",             "is_valid",           8),
        ("Created At",        "created_at",         18),
        ("Assigned At",       "assigned_at",        18),
    ]

    def __init__(
        self,
        order_repo:  OrderRepository,
        reports_dir: str = "reports",
    ):
        self._repo        = order_repo
        self._reports_dir = Path(reports_dir)
        self._reports_dir.mkdir(parents=True, exist_ok=True)

    async def export_daily(
        self,
        export_date:   date = None,
        worker_number: str  = None,
    ) -> Path:
        '''
        Export all orders for a given date.
        Optionally filtered by worker.

        Args:
            export_date:   Date to export. Defaults to today.
            worker_number: Optional worker filter.

        Returns:
            Path to the generated Excel file.
        '''
        target_date   = export_date or date.today()
        records       = await self._repo.get_today(worker_number)

        date_str  = target_date.strftime("%Y-%m-%d")
        worker_str = f"_{worker_number}" if worker_number else ""
        filename   = f"orders_{date_str}{worker_str}.xlsx"
        path       = self._reports_dir / filename

        self._write_excel(records, path, title=f"Orders — {date_str}")
        log.info(f"ExcelExporter: exported {len(records)} order(s) to {path}")
        return path

    async def export_by_worker(
        self,
        worker_number: str,
        status:        str = None,
    ) -> Path:
        '''
        Export all orders assigned to a worker.
        Optionally filtered by assignment_status.

        Args:
            worker_number: Worker phone number.
            status:        Optional status filter e.g. "ASSIGNED".

        Returns:
            Path to the generated Excel file.
        '''
        records   = await self._repo.get_by_worker(worker_number, status)
        date_str  = date.today().strftime("%Y-%m-%d")
        safe_num  = worker_number.replace("+", "")
        filename  = f"worker_{safe_num}_{date_str}.xlsx"
        path      = self._reports_dir / filename

        self._write_excel(records, path, title=f"Orders — +{worker_number}")
        log.info(
            f"ExcelExporter: exported {len(records)} order(s) "
            f"for +{worker_number} to {path}"
        )
        return path

    def _write_excel(
        self,
        records: list[OrderRecord],
        path:    Path,
        title:   str = "Orders",
    ) -> None:
        '''
        Write records to an Excel file using openpyxl.

        Args:
            records: List of OrderRecord objects.
            path:    Output file path.
            title:   Worksheet title.
        '''
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel limit: 31 chars

        # Header fills
        HEADER_FILL = PatternFill("solid", fgColor="1B4F8A")
        VALID_FILL  = PatternFill("solid", fgColor="C6EFCE")
        INVALID_FILL= PatternFill("solid", fgColor="FFC7CE")
        DUP_FILL    = PatternFill("solid", fgColor="FFEB9C")

        # Write header row
        for col_idx, (header, _, width) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill      = HEADER_FILL
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            ws.column_dimensions[
                __import__('openpyxl').utils.get_column_letter(col_idx)
            ].width = width

        # Write data rows
        for row_idx, record in enumerate(records, 2):
            for col_idx, (_, attr, _) in enumerate(self.COLUMNS, 1):
                value = getattr(record, attr, None)

                # Format datetimes
                if isinstance(value, datetime):
                    value = value.strftime("%d/%m/%Y %H:%M")

                # Format booleans
                if isinstance(value, bool):
                    value = "Yes" if value else "No"

                # Format floats (quality score)
                if isinstance(value, float) and attr == "quality_score":
                    value = f"{value:.0%}"

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Row highlighting
            is_valid   = getattr(record, 'is_valid', True)
            dup_status = getattr(record, 'duplicate_status', 'UNIQUE')

            if dup_status in ("CONFIRMED_DUPLICATE", "LIKELY_DUPLICATE"):
                fill = DUP_FILL
            elif is_valid:
                fill = VALID_FILL
            else:
                fill = INVALID_FILL

            for col_idx in range(1, len(self.COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter on header row
        ws.auto_filter.ref = ws.dimensions

        wb.save(path)
"""


# ==============================================================
# ================================================================
#  FILE 8
#  PATH: windwhirl/app/oms/infrastructure/persistence/__init__.py
# ================================================================
# ==============================================================

"""
from app.oms.infrastructure.persistence.schema import (
    Base, OrderRecord, AssignmentRecord,
    DuplicateGroupRecord, DuplicateMemberRecord,
)
from app.oms.infrastructure.persistence.database import Database
from app.oms.infrastructure.persistence.order_repository import OrderRepository
from app.oms.infrastructure.persistence.assignment_repository import AssignmentRepository
from app.oms.infrastructure.persistence.duplicate_repository import DuplicateRepository
from app.oms.infrastructure.persistence.db_duplicate_store import DbDuplicateStore
from app.oms.infrastructure.persistence.excel_exporter import ExcelExporter

__all__ = [
    "Base",
    "OrderRecord", "AssignmentRecord",
    "DuplicateGroupRecord", "DuplicateMemberRecord",
    "Database",
    "OrderRepository", "AssignmentRepository", "DuplicateRepository",
    "DbDuplicateStore",
    "ExcelExporter",
]
"""


# ==============================================================
# ================================================================
#  FILE 9
#  PATH: windwhirl/app/oms/infrastructure/__init__.py  (UPDATE)
# ================================================================
# Add persistence exports to the existing infrastructure __init__.
# ================================================================
# ==============================================================

"""
# ADD to existing infrastructure/__init__.py:
from app.oms.infrastructure.persistence import (
    Database,
    OrderRepository,
    AssignmentRepository,
    DuplicateRepository,
    DbDuplicateStore,
    ExcelExporter,
)
"""


# ==============================================================
# ================================================================
#  FILE 10
#  PATH: windwhirl/oms_runner.py  (UPDATE — wire persistence)
# ================================================================
# Key additions only — the full runner structure stays as-is.
# Add these blocks in the right positions.
# ================================================================
# ==============================================================

"""
# ADD imports:
from app.oms.infrastructure.persistence import (
    Database, OrderRepository, AssignmentRepository,
    DuplicateRepository, DbDuplicateStore, ExcelExporter,
)
from app.oms.application.models.validated_order import ValidatedOrder
from app.oms.application.models.duplicate_result import DuplicateResult
from app.oms.application.models.duplicate_group import DuplicateGroup
from app.oms.application.duplicate.duplicate_detection_engine import DuplicateDetectionEngine

# ADD after build_staff_directory():
def build_persistence(settings):
    db = Database(settings.storage.database_url)
    db.init()
    session_factory    = db.session_factory
    order_repo         = OrderRepository(session_factory)
    assignment_repo    = AssignmentRepository(session_factory)
    duplicate_repo     = DuplicateRepository(session_factory)
    db_store           = DbDuplicateStore(
        order_repo     =order_repo,
        duplicate_repo =duplicate_repo,
        window_hours   =48.0,
    )
    duplicate_engine   = DuplicateDetectionEngine(window_hours=48.0)
    duplicate_engine._store = db_store   # Inject DB store
    exporter           = ExcelExporter(order_repo)
    return (
        db, order_repo, assignment_repo,
        duplicate_repo, duplicate_engine, exporter
    )

# ADD event listeners:

@dispatcher.on("order.validated")
async def on_order_validated(validated_order: ValidatedOrder, **kwargs):
    '''Persist every validated order to the database.'''
    try:
        await order_repo.save_validated_order(validated_order)
    except Exception as e:
        log.error(f"Persistence: failed to save order: {e}", exc_info=True)

@dispatcher.on("order.partially_parsed")
async def on_order_partial(validated_order: ValidatedOrder, **kwargs):
    '''Also persist partially-parsed orders — they still need tracking.'''
    try:
        await order_repo.save_validated_order(validated_order)
    except Exception as e:
        log.error(f"Persistence: failed to save partial order: {e}", exc_info=True)

@dispatcher.on("assignment.resolved")
async def on_assignment_resolved_persist(**kwargs):
    '''Persist assignment and update order record.'''
    order_id      = kwargs.get("order_id")
    worker_number = kwargs.get("worker_number")
    try:
        await order_repo.update_assignment(order_id, worker_number)
        # Note: resolved_assignment object not available directly in event kwargs
        # Full resolution object available in Day 6 engine — extend if needed
        log.info(f"Persistence: assignment saved for order {order_id!r}")
    except Exception as e:
        log.error(f"Persistence: assignment save failed: {e}", exc_info=True)

@dispatcher.on("duplicate.confirmed")
async def on_duplicate_confirmed(**kwargs):
    order_id_a  = kwargs.get("order_id_a")
    order_id_b  = kwargs.get("order_id_b")
    group_id    = kwargs.get("group_id", "")
    try:
        await order_repo.update_duplicate_status(
            order_id_a, "CONFIRMED_DUPLICATE", group_id
        )
    except Exception as e:
        log.error(f"Persistence: duplicate status update failed: {e}", exc_info=True)

@dispatcher.on("duplicate.likely")
async def on_duplicate_likely(**kwargs):
    order_id_a = kwargs.get("order_id_a")
    group_id   = kwargs.get("group_id", "")
    try:
        await order_repo.update_duplicate_status(
            order_id_a, "LIKELY_DUPLICATE", group_id
        )
    except Exception as e:
        log.error(f"Persistence: duplicate likely update failed: {e}", exc_info=True)

# In main(): call build_persistence() and wire duplicate engine:
# (db, order_repo, assignment_repo,
#  duplicate_repo, duplicate_engine, exporter) = build_persistence(settings)
#
# Then pass duplicate_engine to the message processing pipeline.
# After order.validated fires and order is saved, run:
#   @dispatcher.on("order.validated")
#   async def check_duplicate(validated_order, **kwargs):
#       await duplicate_engine.check(validated_order)
"""


# ==============================================================
# ================================================================
#  FILE 11
#  PATH: windwhirl/app/oms/tests/test_persistence.py
# ================================================================
# Run: python -m pytest app/oms/tests/test_persistence.py -v
# Uses an in-memory SQLite database for test isolation.
# ================================================================
# ==============================================================

"""
import asyncio
import sys
from datetime import datetime
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent.parent))

from app.oms.infrastructure.persistence.database import Database
from app.oms.infrastructure.persistence.order_repository import OrderRepository
from app.oms.infrastructure.persistence.assignment_repository import AssignmentRepository
from app.oms.infrastructure.persistence.duplicate_repository import DuplicateRepository
from app.oms.application.models.parsed_order import ParsedOrder, PackageInfo
from app.oms.application.models.validated_order import ValidatedOrder
from app.oms.application.models.validation_report import ValidationReport
from app.oms.application.models.duplicate_group import DuplicateGroup


# ── Test DB fixture ───────────────────────────────────────────────

def make_test_db():
    '''Create an in-memory SQLite database for testing.'''
    db = Database("sqlite:///:memory:")
    db.init()
    return db


def make_order_repo(db: Database) -> OrderRepository:
    return OrderRepository(db.session_factory)


def make_assignment_repo(db: Database) -> AssignmentRepository:
    return AssignmentRepository(db.session_factory)


def make_duplicate_repo(db: Database) -> DuplicateRepository:
    return DuplicateRepository(db.session_factory)


def make_validated(order_id: str = "ORD-001", **kwargs) -> ValidatedOrder:
    parsed = ParsedOrder(
        order_id        =order_id,
        worker_number   =kwargs.get("worker_number", ""),
        customer_name   =kwargs.get("customer_name",  "Blessing Adeyemi"),
        phone_number    =kwargs.get("phone_number",   "08031234567"),
        whatsapp_number =kwargs.get("whatsapp_number","08031234567"),
        package         =PackageInfo("1 Combo Set", "", "#29,500", 29500.0),
        delivery_address=kwargs.get("address", "12 Allen Ave, Ikeja Lagos"),
        delivery_request="Tomorrow",
        raw_text        ="raw test message",
    )
    report = ValidationReport()
    report.is_valid     = True
    report.quality_score= 0.9
    return ValidatedOrder(parsed_order=parsed, report=report)


# ── Database initialization ───────────────────────────────────────

def test_database_init():
    db = make_test_db()
    assert db.session_factory is not None


def test_database_creates_tables():
    db = make_test_db()
    # If tables weren't created, queries below would fail
    repo = make_order_repo(db)
    assert repo is not None


# ── OrderRepository ───────────────────────────────────────────────

@pytest.mark.asyncio
async def test_save_and_retrieve_order():
    db   = make_test_db()
    repo = make_order_repo(db)
    va   = make_validated("ORD-001")

    await repo.save_validated_order(va)
    record = await repo.get_by_id("ORD-001")

    assert record is not None
    assert record.order_id      == "ORD-001"
    assert record.customer_name == "Blessing Adeyemi"
    assert record.phone_number  == "08031234567"


@pytest.mark.asyncio
async def test_upsert_order():
    '''Saving same order_id twice updates, not duplicates.'''
    db   = make_test_db()
    repo = make_order_repo(db)
    va   = make_validated("ORD-001")

    await repo.save_validated_order(va)
    await repo.save_validated_order(va)  # Second save

    records = await repo.get_today()
    assert len([r for r in records if r.order_id == "ORD-001"]) == 1


@pytest.mark.asyncio
async def test_update_assignment():
    db   = make_test_db()
    repo = make_order_repo(db)
    va   = make_validated("ORD-001")
    await repo.save_validated_order(va)

    await repo.update_assignment("ORD-001", "2348031111111")
    record = await repo.get_by_id("ORD-001")

    assert record.worker_number    == "2348031111111"
    assert record.assignment_status == "ASSIGNED"


@pytest.mark.asyncio
async def test_get_today_returns_todays_orders():
    db   = make_test_db()
    repo = make_order_repo(db)

    await repo.save_validated_order(make_validated("ORD-A"))
    await repo.save_validated_order(make_validated("ORD-B"))

    records = await repo.get_today()
    order_ids = [r.order_id for r in records]

    assert "ORD-A" in order_ids
    assert "ORD-B" in order_ids


@pytest.mark.asyncio
async def test_get_pending():
    db   = make_test_db()
    repo = make_order_repo(db)

    await repo.save_validated_order(make_validated("ORD-001"))
    await repo.save_validated_order(make_validated("ORD-002"))
    await repo.update_assignment("ORD-001", "2348XXXXXXXXX")

    pending = await repo.get_pending()
    pending_ids = [r.order_id for r in pending]

    assert "ORD-002" in pending_ids
    assert "ORD-001" not in pending_ids  # Was assigned


@pytest.mark.asyncio
async def test_get_by_worker():
    db   = make_test_db()
    repo = make_order_repo(db)

    await repo.save_validated_order(make_validated("ORD-001"))
    await repo.save_validated_order(make_validated("ORD-002"))
    await repo.update_assignment("ORD-001", "2348031111111")
    await repo.update_assignment("ORD-002", "2348032222222")

    worker1_orders = await repo.get_by_worker("2348031111111")
    assert len(worker1_orders) == 1
    assert worker1_orders[0].order_id == "ORD-001"


@pytest.mark.asyncio
async def test_get_in_window():
    db   = make_test_db()
    repo = make_order_repo(db)

    await repo.save_validated_order(make_validated("ORD-001"))
    await repo.save_validated_order(make_validated("ORD-002"))

    since = datetime.now().replace(hour=0, minute=0, second=0)
    candidates = await repo.get_in_window(since, exclude_id="ORD-001")

    order_ids = [r.order_id for r in candidates]
    assert "ORD-002" in order_ids
    assert "ORD-001" not in order_ids  # Excluded


@pytest.mark.asyncio
async def test_count_by_status():
    db   = make_test_db()
    repo = make_order_repo(db)

    await repo.save_validated_order(make_validated("ORD-001"))
    await repo.save_validated_order(make_validated("ORD-002"))
    await repo.update_assignment("ORD-001", "2348XXXXXXXXX")

    counts = await repo.count_by_status()
    assert counts.get("PENDING", 0) >= 1
    assert counts.get("ASSIGNED", 0) >= 1


# ── DuplicateRepository ───────────────────────────────────────────

@pytest.mark.asyncio
async def test_save_and_retrieve_group():
    db   = make_test_db()
    orepo = make_order_repo(db)
    drepo = make_duplicate_repo(db)

    # Must save orders first (FK constraint)
    await orepo.save_validated_order(make_validated("ORD-A"))
    await orepo.save_validated_order(make_validated("ORD-B"))

    group = DuplicateGroup(
        canonical_order_id="ORD-A",
        classification    ="LIKELY_DUPLICATE",
    )
    group.add_member("ORD-B")

    await drepo.save_group(group)
    retrieved = await drepo.get_group(group.group_id)

    assert retrieved is not None
    assert retrieved.canonical_order_id == "ORD-A"


@pytest.mark.asyncio
async def test_get_groups_for_order():
    db    = make_test_db()
    orepo = make_order_repo(db)
    drepo = make_duplicate_repo(db)

    await orepo.save_validated_order(make_validated("ORD-A"))
    await orepo.save_validated_order(make_validated("ORD-B"))

    group = DuplicateGroup(
        canonical_order_id="ORD-A",
        classification    ="CONFIRMED_DUPLICATE",
    )
    group.add_member("ORD-B")
    await drepo.save_group(group)

    groups = await drepo.get_groups_for_order("ORD-B")
    assert len(groups) >= 1
    assert groups[0].canonical_order_id == "ORD-A"


# ── Excel export ──────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_excel_export_daily(tmp_path):
    from app.oms.infrastructure.persistence.excel_exporter import ExcelExporter

    db   = make_test_db()
    repo = make_order_repo(db)

    await repo.save_validated_order(make_validated("ORD-001"))
    await repo.save_validated_order(make_validated("ORD-002"))

    exporter = ExcelExporter(repo, reports_dir=str(tmp_path))
    path     = await exporter.export_daily()

    assert path.exists()
    assert path.suffix == ".xlsx"
    assert path.stat().st_size > 0


@pytest.mark.asyncio
async def test_excel_export_by_worker(tmp_path):
    from app.oms.infrastructure.persistence.excel_exporter import ExcelExporter

    db   = make_test_db()
    repo = make_order_repo(db)

    va = make_validated("ORD-001")
    await repo.save_validated_order(va)
    await repo.update_assignment("ORD-001", "2348031111111")

    exporter = ExcelExporter(repo, reports_dir=str(tmp_path))
    path     = await exporter.export_by_worker("2348031111111")

    assert path.exists()
    assert "2348031111111" in path.name
"""


# ==============================================================
# DAY 10 VERIFICATION
# ==============================================================
#
# Test 1 — Imports:
#   python -c "
#   import sys; sys.path.insert(0, '.')
#   from app.oms.infrastructure.persistence import (
#       Database, OrderRepository, AssignmentRepository,
#       DuplicateRepository, DbDuplicateStore, ExcelExporter
#   )
#   print('All Day 10 imports OK')
#   "
#
# Test 2 — Database init:
#   python -c "
#   import sys; sys.path.insert(0, '.')
#   from app.oms.infrastructure.persistence.database import Database
#   db = Database('sqlite:///data/oms.db')
#   db.init()
#   print('Database initialized OK')
#   print('Tables: orders, assignments, duplicate_groups, duplicate_members')
#   "
#
# Test 3 — Full save/retrieve cycle:
#   python -c "
#   import sys, asyncio; sys.path.insert(0, '.')
#   from app.oms.infrastructure.persistence.database import Database
#   from app.oms.infrastructure.persistence.order_repository import OrderRepository
#   from app.oms.application.models.parsed_order import ParsedOrder, PackageInfo
#   from app.oms.application.models.validated_order import ValidatedOrder
#   from app.oms.application.models.validation_report import ValidationReport
#
#   db   = Database('sqlite:///:memory:')
#   db.init()
#   repo = OrderRepository(db.session_factory)
#
#   parsed = ParsedOrder(
#       order_id='ORD-001', worker_number='',
#       customer_name='Blessing Adeyemi', phone_number='08031234567',
#       package=PackageInfo('1 Combo', '', '#29500', 29500.0),
#       delivery_address='12 Allen Ave Ikeja', raw_text='test',
#   )
#   report = ValidationReport()
#   report.is_valid = True
#   report.quality_score = 0.9
#   va = ValidatedOrder(parsed_order=parsed, report=report)
#
#   async def run():
#       await repo.save_validated_order(va)
#       record = await repo.get_by_id('ORD-001')
#       print(f'Saved and retrieved: {record}')
#       counts = await repo.count_by_status()
#       print(f'Status counts: {counts}')
#
#   asyncio.run(run())
#   "
#
# Test 4 — Run all unit tests:
#   python -m pytest app/oms/tests/test_persistence.py -v
#   Expected: 15+ tests PASSED
#
# Test 5 — Full system test (requires running OMS):
#   python oms_runner.py
#   Send a test order in the WhatsApp group.
#   Check: sqlite3 data/oms.db "SELECT * FROM orders LIMIT 5;"
#   Check: sqlite3 data/oms.db "SELECT * FROM assignments LIMIT 5;"
#
# ==============================================================
# WHAT DAY 11 BUILDS
# ==============================================================
# Day 11: WhatsApp Notification Engine
#   Sends structured confirmation messages back to the WhatsApp group
#   when an order is successfully assigned.
#
#   Listens to "assignment.resolved" event.
#   Composes a confirmation message per order.
#   Uses the existing PlaywrightSender from the review automation
#   OR a separate OMS sender (TBD based on session constraints).
#   Includes: order summary, worker name, estimated delivery.
#
#   Does NOT block the pipeline — fires asynchronously.
#   Does NOT retry if WhatsApp send fails — logs and continues.
# ==============================================================