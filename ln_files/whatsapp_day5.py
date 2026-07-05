# ==============================================================
# WHATSAPP REVIEW AUTOMATION — DAY 5 BUILD
# ==============================================================
# FILES IN THIS DOCUMENT:
#   FILE 15 → apps/core/lib/utils/excel_reporter.py  (new file)
#   FILE 16 → apps/core/lib/utils/reporter.py        (replace existing)
#   FILE 17 → apps/core/lib/utils/playwright_sender.py (add one method)
#
# WHAT YOU ARE BUILDING TODAY:
#
#   FILE 15 — ExcelReporter (new class)
#     Reads your original customers.xlsx, adds one new column
#     "Status/Comment" after the WhatsApp Number column, fills it
#     with the send result for each customer, saves the file to
#     reports/ with today's date in the filename.
#
#   FILE 16 — Updated Reporter
#     Replaces the existing reporter.py. Adds the call to
#     ExcelReporter and the WhatsApp file-send logic. Removes
#     the email logic entirely as requested.
#
#   FILE 17 — One new method in PlaywrightSender
#     send_file_to_number() — sends a file (the Excel report)
#     to your personal WhatsApp number as a document attachment.
#     Retries twice. If both fail, saves to reports/ only.
#
# HOW IT FLOWS:
#   End of last session
#     → ExcelReporter builds the Excel file with Status/Comment
#     → Saves to reports/send_report_YYYY-MM-DD.xlsx
#     → PlaywrightSender.send_file_to_number() sends it to your
#       personal WhatsApp
#     → If WhatsApp send fails twice → file stays in reports/ only
#     → Text summary report still saved to reports/ as .txt
#
# STATUS/COMMENT COLUMN VALUES:
#   "True / Sent"
#   "False / Not sent — Phone not registered on WhatsApp"
#   "False / Not sent — Send timeout after 2 attempts"
#   "False / Not sent — Invalid phone number"
#   "False / Not sent — Pending (not yet attempted)"
#   "False / Not sent — [actual error message from system]"
#
# CONFIG YOU NEED TO ADD:
#   In apps/core/config.py → CONFIG dict, add:
#     "personal_whatsapp": "2348XXXXXXXXX"  ← your personal number
#
# FOLDER STRUCTURE AFTER TODAY:
#   apps/
#   ├── reports/
#   │   ├── send_report_2025-06-28.xlsx   ← generated Excel
#   │   └── daily_report_2025-06-28.txt   ← text summary
#   └── core/lib/utils/
#       ├── excel_reporter.py              ← FILE 15 (new)
#       ├── reporter.py                    ← FILE 16 (replace)
#       └── playwright_sender.py           ← FILE 17 (add method)
# ==============================================================


# ==============================================================
# FIRST: Add this to your CONFIG dict in apps/core/config.py
# ==============================================================
# Find the CONFIG dict and add this one line anywhere inside it:
#
#   "personal_whatsapp": "2348XXXXXXXXX",  # EDIT ME — your number
#
# Then add it to AppConfig.__init__() as:
#   self.personal_whatsapp = str(raw.get("personal_whatsapp", ""))
#
# And add a helper method to AppConfig:
#   def has_personal_whatsapp(self) -> bool:
#       return bool(self.personal_whatsapp.strip())
# ==============================================================


# ==============================================================
# ================================================================
#  FILE 15
#  PATH:  apps/core/lib/utils/excel_reporter.py
#  TYPE:  Python file — NEW file, create it fresh
# ================================================================
# PURPOSE:
#   Reads the original customers.xlsx, pulls each customer's
#   send result from the database, inserts one new column
#   "Status/Comment" right after the WhatsApp Number column,
#   and saves the result as a dated Excel file in reports/.
#
# OUTPUT FILE:
#   reports/send_report_YYYY-MM-DD.xlsx
#
# THE NEW COLUMN:
#   Name:     "Status/Comment"
#   Position: Immediately after "WhatsApp Number" column
#   Values:
#     "True / Sent"
#     "False / Not sent — [reason]"
#
# WHY READ THE ORIGINAL EXCEL:
#   We keep ALL the original columns intact so you can open
#   the file and immediately see name, phone, product AND
#   send status side by side. No data is removed or reordered.
#   Only one column is added.
#
# FUTURE / FASTAPI NOTE:
#   GET /reports/excel/today → returns the Excel file as a
#   file response for download from the Next.js dashboard.
#   No changes to this class needed for that.
# ================================================================
# ==============================================================

# ── COPY EVERYTHING BELOW THIS LINE INTO:
# ── apps/core/lib/utils/excel_reporter.py ─────────────────────

import logging
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelReporter:
    """
    Generates a status Excel report from the original customers.xlsx.

    Takes the original Excel file as a base, adds one column
    "Status/Comment" after the WhatsApp Number column, fills it
    with the send result for each customer from the database,
    and saves to reports/send_report_YYYY-MM-DD.xlsx.

    Usage:
        reporter = ExcelReporter(cfg)
        path = reporter.generate(db)
        # path → Path("reports/send_report_2025-06-28.xlsx")
    """

    # ── Cell colours for visual scanning ──────────────────────
    # Green fill for sent, red fill for not sent
    # Makes the report scannable at a glance without reading
    GREEN_FILL = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )
    RED_FILL = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )
    HEADER_FILL = PatternFill(
        start_color="1B4F8A",
        end_color="1B4F8A",
        fill_type="solid"
    )

    def __init__(self, cfg):
        """
        Args:
            cfg: AppConfig instance — provides excel_path() and country_code
        """
        self._cfg = cfg
        self._log = logging.getLogger(self.__class__.__name__)

    def generate(self, db) -> Path:
        """
        Build the status Excel report and save it to reports/.

        Flow:
          1. Read original customers.xlsx into pandas
          2. Pull all send statuses from database
          3. Match each Excel row to its DB status by Order ID
          4. Build "Status/Comment" value per row
          5. Insert new column after WhatsApp Number
          6. Apply colour coding (green=sent, red=not sent)
          7. Save to reports/send_report_YYYY-MM-DD.xlsx

        Args:
            db: Database instance

        Returns:
            Path to the saved Excel file.

        Raises:
            FileNotFoundError if original Excel is missing.
        """
        excel_path = self._cfg.excel_path()
        if not excel_path.exists():
            raise FileNotFoundError(
                f"Original Excel not found: {excel_path}\n"
                f"Cannot generate report without source file."
            )

        self._log.info(f"Building Excel report from: {excel_path}")

        # ── Step 1: Read original Excel ────────────────────────
        # Keep all columns exactly as they are in the original.
        # dtype=str prevents pandas from mangling phone numbers.
        df = pd.read_excel(excel_path, dtype=str, engine="openpyxl")

        # ── Step 2: Get all send statuses from DB ──────────────
        # Returns dict: { order_id: {"status": ..., "error": ...} }
        status_map = db.get_all_statuses()

        # ── Step 3: Build Status/Comment value for each row ────
        status_comments = []

        for _, row in df.iterrows():
            order_id = str(row.get("Order ID", "") or "").strip()

            if not order_id or order_id.lower() == "nan":
                # Row has no Order ID — cannot match to DB
                status_comments.append("False / Not sent — No Order ID found")
                continue

            record = status_map.get(order_id)

            if record is None:
                # Order ID exists in Excel but not in database
                # (row was skipped during import — wrong product etc.)
                status_comments.append(
                    "False / Not sent — Not in target product filter"
                )
                continue

            status        = record["status"]
            error_message = record.get("error_message", "") or ""

            # ── Map DB status to human-readable Status/Comment ─
            if status == "SENT":
                status_comments.append("True / Sent")

            elif status == "INVALID_NUMBER":
                status_comments.append(
                    "False / Not sent — Phone not registered on WhatsApp"
                )

            elif status == "FAILED_FINAL":
                reason = error_message if error_message else "Send failed after 2 attempts"
                status_comments.append(f"False / Not sent — {reason}")

            elif status == "FAILED":
                reason = error_message if error_message else "Send failed — eligible for retry"
                status_comments.append(f"False / Not sent — {reason}")

            elif status == "PENDING":
                status_comments.append(
                    "False / Not sent — Pending (not yet attempted)"
                )

            elif status == "INVALID_PHONE":
                status_comments.append(
                    "False / Not sent — Phone number could not be normalized"
                )

            else:
                # Catch-all for any unexpected status value
                status_comments.append(
                    f"False / Not sent — {status}"
                )

        # ── Step 4: Insert Status/Comment column ───────────────
        # Find position of WhatsApp Number column
        # Insert our new column immediately after it
        columns = list(df.columns)

        if "WhatsApp Number" in columns:
            insert_position = columns.index("WhatsApp Number") + 1
        else:
            # WhatsApp Number column not found — append at end
            self._log.warning(
                "WhatsApp Number column not found in Excel. "
                "Status/Comment will be appended at the end."
            )
            insert_position = len(columns)

        # Insert the new column into the dataframe
        df.insert(
            loc=insert_position,
            column="Status/Comment",
            value=status_comments
        )

        # ── Step 5: Save to reports/ ───────────────────────────
        Path("reports").mkdir(exist_ok=True)
        today_str   = date.today().strftime("%Y-%m-%d")
        output_path = Path("reports") / f"send_report_{today_str}.xlsx"

        # Save via pandas first (handles data correctly)
        df.to_excel(output_path, index=False, engine="openpyxl")

        # ── Step 6: Apply colour coding with openpyxl ──────────
        # Re-open the saved file to apply styling
        # pandas doesn't support cell-level styling directly
        self._apply_styling(output_path, insert_position + 1)
        # +1 because openpyxl columns are 1-indexed

        self._log.info(
            f"Excel report saved: {output_path} "
            f"({len(df)} rows, {len(df.columns)} columns)"
        )

        return output_path

    def _apply_styling(self, file_path: Path, status_col_index: int):
        """
        Apply visual styling to the saved Excel file.

        - Header row: dark blue background, white bold text
        - Status/Comment column header: same dark blue
        - "True / Sent" cells: green background
        - "False / Not sent" cells: red background
        - Auto-width on Status/Comment column

        Args:
            file_path:        Path to the Excel file to style.
            status_col_index: 1-based column index of Status/Comment.
        """
        try:
            wb   = load_workbook(file_path)
            ws   = wb.active
            col  = status_col_index
            col_letter = get_column_letter(col)

            # Style the header row (row 1)
            for cell in ws[1]:
                cell.fill      = self.HEADER_FILL
                cell.font      = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

            # Style each data cell in Status/Comment column
            for row_num in range(2, ws.max_row + 1):
                cell  = ws[f"{col_letter}{row_num}"]
                value = str(cell.value or "")

                if value.startswith("True"):
                    cell.fill = self.GREEN_FILL
                elif value.startswith("False"):
                    cell.fill = self.RED_FILL

                # Wrap text so long error messages are readable
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )

            # Auto-width the Status/Comment column
            # (set to 45 chars — enough for most error messages)
            ws.column_dimensions[col_letter].width = 45

            # Freeze the header row so it stays visible when scrolling
            ws.freeze_panes = "A2"

            wb.save(file_path)
            self._log.debug(f"Styling applied to: {file_path}")

        except Exception as e:
            # Styling failure must not prevent the report from saving
            # The data is already saved — styling is cosmetic only
            self._log.warning(f"Could not apply Excel styling: {e}")

# ── END OF FILE 15 ────────────────────────────────────────────


# ==============================================================
# ================================================================
#  FILE 16
#  PATH:  apps/core/lib/utils/reporter.py
#  TYPE:  Python file — REPLACE your existing reporter.py
# ================================================================
# PURPOSE:
#   Replaces the existing reporter. Removes email entirely.
#   Adds:
#     1. Call to ExcelReporter to generate the Excel file
#     2. Call to PlaywrightSender to send the file to your
#        personal WhatsApp number
#     3. Retry logic: tries WhatsApp send twice, then gives up
#        and leaves the file in reports/ only
#     4. Still generates the text summary report to reports/
#
# WHAT STAYS THE SAME:
#   generate_report(db) → still produces the .txt summary
#
# WHAT IS NEW:
#   generate_excel_report(db) → produces the .xlsx file
#   send_report_via_whatsapp(sender, cfg, file_path) → sends it
# ================================================================
# ==============================================================

# ── COPY EVERYTHING BELOW THIS LINE INTO:
# ── apps/core/lib/utils/reporter.py ───────────────────────────

import asyncio
import logging
from datetime import datetime, date
from pathlib import Path

logger = logging.getLogger(__name__)


class Reporter:
    """
    Generates daily reports and delivers them.

    Produces two report files in reports/:
      1. send_report_YYYY-MM-DD.xlsx  — full customer list with status
      2. daily_report_YYYY-MM-DD.txt  — summary counts

    Delivery:
      Sends the Excel file to personal WhatsApp number.
      Retries twice on failure.
      If both attempts fail: file stays in reports/ only.
      No email. No other delivery method.

    Usage:
        reporter = Reporter(cfg)
        await reporter.run_end_of_day(db, sender)
    """

    def __init__(self, cfg):
        """
        Args:
            cfg: AppConfig instance
        """
        self._cfg = cfg
        self._log = logging.getLogger(self.__class__.__name__)

    async def run_end_of_day(self, db, sender):
        """
        Master end-of-day method. Called by Scheduler after last session.

        Flow:
          1. Generate text summary → reports/daily_report_YYYY-MM-DD.txt
          2. Generate Excel report → reports/send_report_YYYY-MM-DD.xlsx
          3. Send Excel to personal WhatsApp (2 attempts max)
          4. Log outcome

        Args:
            db:     Database instance
            sender: PlaywrightSender instance (already connected)
        """
        self._log.info("Running end-of-day report generation...")

        # ── Step 1: Text summary ───────────────────────────────
        try:
            summary_text = self.generate_text_report(db)
            print("\n" + summary_text)
        except Exception as e:
            self._log.error(f"Text report failed: {e}", exc_info=True)
            summary_text = None

        # ── Step 2: Excel report ───────────────────────────────
        excel_path = None
        try:
            from apps.core.lib.utils.excel_reporter import ExcelReporter
            excel_reporter = ExcelReporter(self._cfg)
            excel_path     = excel_reporter.generate(db)
            self._log.info(f"Excel report ready: {excel_path}")
        except Exception as e:
            self._log.error(f"Excel report failed: {e}", exc_info=True)

        # ── Step 3: Send Excel via WhatsApp ────────────────────
        if excel_path and self._cfg.has_personal_whatsapp():
            await self._send_via_whatsapp(sender, excel_path)
        elif excel_path:
            self._log.info(
                "personal_whatsapp not configured — "
                f"Excel report saved to {excel_path} only."
            )
        else:
            self._log.warning(
                "Excel report was not generated — nothing to send."
            )

    async def _send_via_whatsapp(self, sender, file_path: Path):
        """
        Send the Excel report to personal WhatsApp number.
        Retries once on failure. Gives up after 2 total attempts.

        The file always stays in reports/ regardless of outcome.
        WhatsApp send is best-effort — failure is logged but
        does not raise an exception.

        Args:
            sender:    PlaywrightSender instance (already connected)
            file_path: Path to the Excel file to send
        """
        phone    = self._cfg.personal_whatsapp
        caption  = (
            f"📊 Nabeau Store — Daily Send Report\n"
            f"Date: {date.today().strftime('%d %B %Y')}\n"
            f"File: {file_path.name}"
        )

        max_attempts = 2

        for attempt in range(1, max_attempts + 1):
            self._log.info(
                f"Sending report to personal WhatsApp "
                f"+{phone} (attempt {attempt}/{max_attempts})..."
            )
            try:
                result = await sender.send_file_to_number(
                    phone=phone,
                    file_path=str(file_path),
                    caption=caption,
                    order_id=f"REPORT_{date.today().strftime('%Y%m%d')}"
                )

                if result.success:
                    self._log.info(
                        f"✅ Report sent to WhatsApp +{phone} successfully."
                    )
                    return   # Success — exit retry loop

                # Send returned a result but with failure status
                self._log.warning(
                    f"Attempt {attempt} failed: {result.error_message}"
                )

            except Exception as e:
                self._log.warning(
                    f"Attempt {attempt} exception: {e}"
                )

            # Wait before retry (only if there is a next attempt)
            if attempt < max_attempts:
                self._log.info("Waiting 30s before retry...")
                await asyncio.sleep(30)

        # Both attempts failed
        self._log.warning(
            f"WhatsApp report delivery failed after {max_attempts} attempts.\n"
            f"Report is saved locally at: {file_path}\n"
            f"Open it manually from the reports/ folder."
        )

    def generate_text_report(self, db) -> str:
        """
        Generate plain text summary report and save to reports/.
        Returns the report string for printing to console.

        Args:
            db: Database instance

        Returns:
            Report as a plain text string.
        """
        summary   = db.get_daily_summary()
        today_str = date.today().strftime("%Y-%m-%d")

        lines = [
            f"Nabeau Store — WhatsApp Send Report — {today_str}",
            "=" * 52,
            "",
            "SUMMARY",
            f"  Sent:             {summary.get('SENT', 0)}",
            f"  Failed:           {summary.get('FAILED', 0)}",
            f"  Failed (final):   {summary.get('FAILED_FINAL', 0)}",
            f"  Invalid number:   {summary.get('INVALID_NUMBER', 0)}",
            f"  Pending:          {summary.get('PENDING', 0)}",
            "",
            "TEMPLATE PERFORMANCE (today)",
            f"  Template A sent:  {summary.get('template_A', 0)}",
            f"  Template B sent:  {summary.get('template_B', 0)}",
            "",
        ]

        # Failed customer details
        failed = summary.get("failed_details", [])
        if failed:
            lines.append("FAILED — run --reset-failed to retry tomorrow")
            lines.append("-" * 40)
            for item in failed:
                lines.append(
                    f"  {item['name']}  |  "
                    f"+{item['phone']}  |  "
                    f"{item['error']}"
                )
        else:
            lines.append("FAILED: None ✅")

        lines.append("")

        # Invalid number details
        invalid = summary.get("invalid_details", [])
        if invalid:
            lines.append("INVALID NUMBERS — not on WhatsApp")
            lines.append("-" * 40)
            for item in invalid:
                lines.append(f"  {item['name']}  |  {item['phone']}")
        else:
            lines.append("INVALID NUMBERS: None ✅")

        lines += [
            "",
            "=" * 52,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        ]

        report_text = "\n".join(lines)

        # Save text report
        Path("reports").mkdir(exist_ok=True)
        report_path = Path("reports") / f"daily_report_{today_str}.txt"
        report_path.write_text(report_text, encoding="utf-8")
        self._log.info(f"Text report saved: {report_path}")

        return report_text

    # ── Keep this for backward compatibility with CLI --report ─
    def generate_report(self, db) -> str:
        """Alias for generate_text_report. Used by --report CLI command."""
        return self.generate_text_report(db)

# ── END OF FILE 16 ────────────────────────────────────────────


# ==============================================================
# ================================================================
#  FILE 17
#  PATH:  apps/core/lib/utils/playwright_sender.py
#  TYPE:  Python file — ADD one method to existing class
# ================================================================
# PURPOSE:
#   Add send_file_to_number() method to your existing
#   PlaywrightSender class. Do NOT replace the whole file.
#   Find the class in your existing playwright_sender.py and
#   add this method at the bottom of the class, before the
#   final comment line.
#
# WHERE TO ADD IT:
#   Inside the PlaywrightSender class, after the send_image()
#   method and before the class ends.
#
# HOW IT WORKS:
#   Opens the chat with your personal number.
#   Clicks the attachment button.
#   Selects "Document" (not image) to send the Excel file.
#   Uploads the .xlsx file.
#   Sends with a caption.
#   Returns SendResult like all other send methods.
# ================================================================
# ==============================================================

# ── ADD THIS METHOD INSIDE PlaywrightSender class in:
# ── apps/core/lib/utils/playwright_sender.py ──────────────────
#
# Find the line that says:  # ── END OF FILE 11 ────
# Add this entire method ABOVE that line, still inside the class

"""
    async def send_file_to_number(
        self,
        phone:     str,
        file_path: str,
        caption:   str,
        order_id:  str
    ) -> "SendResult":
        \"\"\"
        Send a file (Excel report) to a WhatsApp number as a document.
        Used by Reporter to deliver the daily Excel report.

        Sends as Document (not Image) so the .xlsx file is received
        as a downloadable file, not converted to an image preview.

        Args:
            phone:     13-digit normalized phone e.g. "2348XXXXXXXXX"
            file_path: Full local path to the file e.g. "reports/send_report_2025-06-28.xlsx"
            caption:   Short description text sent with the file
            order_id:  For logging (use "REPORT_YYYYMMDD" format)

        Returns:
            SendResult with status SENT | FAILED | INVALID_NUMBER
        \"\"\"
        import random
        import asyncio
        from pathlib import Path

        self._log.info(
            f"→ Sending file to +{phone} [{order_id}]\n"
            f"  File: {file_path}"
        )

        # Verify the file actually exists before attempting send
        if not Path(file_path).exists():
            self._log.error(f"File not found: {file_path}")
            return SendResult(
                success=False,
                status="FAILED",
                error_message=f"File not found: {file_path}"
            )

        try:
            await self._rotate_tab()

            # Navigate to the chat — same clean URL as text sends
            # (STEALTH 4: no ?text= parameter)
            await self._page.goto(
                f"https://web.whatsapp.com/send?phone={phone}",
                wait_until="domcontentloaded",
                timeout=20_000
            )

            # Wait for chat to load
            try:
                await self._page.wait_for_selector(
                    self.SEL["msg_input"],
                    timeout=15_000
                )
            except Exception:
                if await self._check_invalid_number():
                    return SendResult(
                        success=False,
                        status="INVALID_NUMBER",
                        error_message="Phone not registered on WhatsApp"
                    )
                raise

            if await self._check_invalid_number():
                return SendResult(
                    success=False,
                    status="INVALID_NUMBER",
                    error_message="Phone not registered on WhatsApp"
                )

            # STEALTH 2: Pre-action pause
            await asyncio.sleep(random.uniform(2, 4))

            # Click the attachment (paperclip) button
            await self._page.click(self.SEL["attach_btn"])
            await asyncio.sleep(random.uniform(0.8, 1.5))

            # ── Select "Document" attachment type ───────────────
            # WhatsApp Web shows options: Photos, Camera, Document, etc.
            # We need Document to send .xlsx as a downloadable file.
            # If we use the image input it will try to render xlsx as image.
            document_input_selector = 'input[accept*="*/*"], input[type="file"]'

            try:
                # Try clicking the Document option in the attachment menu
                doc_option = await self._page.query_selector(
                    'li[data-testid="mi-attach-document"], '
                    'span[data-icon="attach-document"]'
                )
                if doc_option:
                    await doc_option.click()
                    await asyncio.sleep(0.5)
            except Exception:
                pass  # Fall through to direct file input

            # ── Upload the file via file chooser ────────────────
            try:
                async with self._page.expect_file_chooser(
                    timeout=8_000
                ) as fc_info:
                    # Try clicking any visible file input
                    file_inputs = await self._page.query_selector_all(
                        'input[type="file"]'
                    )
                    if file_inputs:
                        await file_inputs[-1].click()
                    else:
                        # Last resort: press Enter on the attachment menu
                        await self._page.keyboard.press("Enter")

                file_chooser = await fc_info.value
                await file_chooser.set_files(file_path)
                self._log.debug(f"  File uploaded to chooser: {file_path}")

            except Exception as e:
                self._log.error(f"  File chooser failed: {e}")
                return SendResult(
                    success=False,
                    status="FAILED",
                    error_message=f"File upload failed: {e}"
                )

            # Wait for file preview to appear in WhatsApp
            await asyncio.sleep(2.0)

            # ── Type the caption ─────────────────────────────────
            # Caption input for documents is different from image caption
            caption_selectors = [
                'div[aria-label="Add a caption"]',
                'div[aria-label="Type a message"]',
                'div[data-tab="10"]',
            ]

            caption_typed = False
            for sel in caption_selectors:
                try:
                    caption_el = await self._page.query_selector(sel)
                    if caption_el:
                        await caption_el.click()
                        await asyncio.sleep(0.5)
                        # Type caption character by character (STEALTH 1)
                        await self._type_human(sel, caption)
                        caption_typed = True
                        break
                except Exception:
                    continue

            if not caption_typed:
                self._log.warning("  Caption input not found — sending without caption")

            await asyncio.sleep(0.5)

            # Send the file
            await self._page.keyboard.press("Enter")
            self._log.debug("  File sent. Waiting for confirmation...")

            # Wait for delivery tick
            try:
                await self._page.wait_for_selector(
                    self.SEL["sent_tick"],
                    timeout=30_000   # Files take longer than text
                )
                self._log.info(f"  ✅ File delivered to +{phone}")
            except Exception:
                self._log.warning(
                    f"  ⚠ Tick timeout for file send to +{phone} — "
                    "likely sent but unconfirmed"
                )

            # Screenshot as proof
            screenshot_path = await self._take_screenshot(order_id)
            self._msgs_on_tab += 1

            return SendResult(
                success=True,
                status="SENT",
                screenshot_path=screenshot_path
            )

        except Exception as e:
            self._log.error(
                f"  ✗ File send failed for +{phone}: {e}",
                exc_info=True
            )
            return SendResult(
                success=False,
                status="FAILED",
                error_message=str(e)
            )
"""

# ── END OF FILE 17 ────────────────────────────────────────────


# ==============================================================
# DATABASE UPDATE NEEDED
# ==============================================================
# Add this method to your Database class in:
# apps/core/db/database.py
#
# The ExcelReporter calls db.get_all_statuses() which doesn't
# exist yet. Add it after the get_stats() method.
# ──────────────────────────────────────────────────────────────

"""
    def get_all_statuses(self) -> dict:
        \"\"\"
        Returns a dict of every customer's send status.
        Used by ExcelReporter to populate the Status/Comment column.

        Returns:
            {
                "ORD001": {
                    "status": "SENT",
                    "error_message": ""
                },
                "ORD002": {
                    "status": "FAILED",
                    "error_message": "Timeout waiting for tick"
                },
                ...
            }
        \"\"\"
        with self._session() as session:
            rows = (
                session.query(Customer, SendLog)
                .join(SendLog, Customer.id == SendLog.customer_id)
                .all()
            )

            result = {}
            for customer, log in rows:
                result[customer.order_id] = {
                    "status":        log.status,
                    "error_message": log.error_message or "",
                }

            return result
"""

# ── END OF DATABASE UPDATE ─────────────────────────────────────


# ==============================================================
# SCHEDULER UPDATE NEEDED
# ==============================================================
# In apps/core/lib/scheduler/scheduler.py
# Find the _end_of_day() method and replace it with this:
# ──────────────────────────────────────────────────────────────

"""
    async def _end_of_day(self):
        \"\"\"
        Run after the final session of the day.
        Generates both reports and sends Excel via WhatsApp.
        \"\"\"
        self._log.info("All sessions complete. Running end-of-day tasks...")

        # Run the full report flow (text + Excel + WhatsApp send)
        await self._reporter.run_end_of_day(self._db, self._sender)

        # Report retry-eligible customers
        retries = self._db.get_retry_eligible()
        if retries:
            self._log.info(
                f"{len(retries)} message(s) eligible for retry. "
                f"Run: python main.py --reset-failed"
            )
        else:
            self._log.info("No messages need retry.")
"""

# ── END OF SCHEDULER UPDATE ───────────────────────────────────


# ==============================================================
# CONFIG UPDATE — Add to CONFIG dict in apps/core/config.py
# ==============================================================

"""
    # ── PERSONAL WHATSAPP ──────────────────────────────────────
    # Your personal WhatsApp number to receive the daily Excel report.
    # Format: country code + number, no + sign, no spaces.
    # Nigeria example: "2348012345678"
    # Leave as "" to skip WhatsApp delivery (report saved locally only).
    "personal_whatsapp": "2348XXXXXXXXX",  # EDIT ME
"""

# And add to AppConfig.__init__():
"""
        self.personal_whatsapp = str(raw.get("personal_whatsapp", "")).strip()
"""

# And add this method to AppConfig class:
"""
    def has_personal_whatsapp(self) -> bool:
        \"\"\"True if a personal WhatsApp number is configured for report delivery.\"\"\"
        return bool(self.personal_whatsapp)
"""

# ── END OF CONFIG UPDATE ───────────────────────────────────────


# ==============================================================
# DAY 5 VERIFICATION
# ==============================================================
#
# After all updates are saved, test the Excel report alone first:
#
#   python -c "
#   import sys
#   from pathlib import Path
#   sys.path.insert(0, str(Path('.').resolve().parent))
#
#   from apps.core.config import AppConfig
#   from apps.core.db.database import Database
#   from apps.core.lib.utils.excel_reporter import ExcelReporter
#
#   cfg = AppConfig()
#   db  = Database(cfg.database_url)
#   db.init()
#
#   reporter = ExcelReporter(cfg)
#   path     = reporter.generate(db)
#   print(f'Report generated: {path}')
#   "
#
#   Expected output:
#     Report generated: reports/send_report_2025-06-28.xlsx
#
#   Open the file in Excel and confirm:
#     ✅ All original columns are intact
#     ✅ "Status/Comment" column appears after WhatsApp Number
#     ✅ Sent rows show green "True / Sent"
#     ✅ Failed/pending rows show red "False / Not sent — [reason]"
#     ✅ First row is frozen (stays visible when scrolling)
#
# Then test full end-of-day:
#   python main.py --report
#
#   Expected:
#     ✅ Text report printed to console
#     ✅ reports/daily_report_YYYY-MM-DD.txt saved
#     ✅ reports/send_report_YYYY-MM-DD.xlsx saved
#     ✅ Excel sent to your personal WhatsApp (if configured)
# ==============================================================
